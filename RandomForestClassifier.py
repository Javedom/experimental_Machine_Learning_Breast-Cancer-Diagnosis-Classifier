import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import pickle
import sys
import argparse

# -------------------------------------------------------------------------
# Imports for file handling and encoding detection
# -------------------------------------------------------------------------
import chardet

# -------------------------------------------------------------------------
# Optimization Import
# -------------------------------------------------------------------------
import optuna

# -------------------------------------------------------------------------
# Scikit-Learn Imports
# -------------------------------------------------------------------------
from sklearn.model_selection import train_test_split, cross_val_score, GridSearchCV, StratifiedKFold
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.metrics import (
    classification_report,
    precision_score,
    recall_score,
    f1_score,
    confusion_matrix,
    ConfusionMatrixDisplay,
    roc_curve,
    auc,
    accuracy_score
)

# -----------------------------------------------------------------------------
# 1) Model Loading Function - Moved to the top for logical flow
# -----------------------------------------------------------------------------
def load_saved_model(model_file='breast_cancer_model.pkl'):
    """
    Loads a previously trained model and its components from a pickle file.
    
    Arguments:
        model_file: Path to the saved model file
        
    Returns:
        model: The trained model
        scaler: StandardScaler fitted on training data
        label_encoder: LabelEncoder for diagnosis labels
        feature_names: List of feature column names
        success: Boolean indicating if loading was successful
    """
    try:
        with open(model_file, 'rb') as f:
            model_data = pickle.load(f)
            
        # Extract components
        model = model_data['model']
        scaler = model_data['scaler']
        label_encoder = model_data['label_encoder']
        feature_names = model_data['feature_names']
        
        print(f"✅ Successfully loaded model from '{model_file}'")
        print(f"Model type: {type(model).__name__}")
        print(f"Number of features: {len(feature_names)}")
        
        return model, scaler, label_encoder, feature_names, True
    except Exception as e:
        print(f"❌ Error loading model: {str(e)}")
        return None, None, None, None, False

# -----------------------------------------------------------------------------
# 2) Data Loading
# -----------------------------------------------------------------------------
def load_data(train_file_base="training_data", predict_file_base="prediction_data"):
    """
    Loads the breast cancer datasets, supporting both CSV and Excel formats.
    
    Automatically detects and loads from either .csv or .xlsx file extensions.
    
    Arguments:
        train_file_base: Base name of training data file (without extension), can be None
        predict_file_base: Base name of prediction data file (without extension), can be None
    
    Returns:
    - train_df: DataFrame with training data (includes diagnosis) or None
    - predict_df: DataFrame with prediction data (or None if not provided)
    - train_format: Format of the loaded training file ('csv' or 'xlsx') or None
    """
    def try_load_file(base_name):
        """Helper function to try loading either CSV or Excel file."""
        if base_name is None:
            return None, None
            
        csv_path = f"{base_name}.csv"
        xlsx_path = f"{base_name}.xlsx"
        
        if os.path.exists(csv_path):
            # Try to detect encoding for CSV
            try:
                with open(csv_path, 'rb') as f:
                    rawdata = f.read(10000)
                    result = chardet.detect(rawdata)
                    encoding = result['encoding']
                df = pd.read_csv(csv_path, encoding=encoding)
                print(f"Successfully loaded {csv_path} with {df.shape[0]} rows and {df.shape[1]} columns")
                return df, 'csv'
            except Exception as e:
                print(f"Error loading CSV: {str(e)}")
                return None, None
                
        elif os.path.exists(xlsx_path):
            try:
                df = pd.read_excel(xlsx_path)
                print(f"Successfully loaded {xlsx_path} with {df.shape[0]} rows and {df.shape[1]} columns")
                return df, 'xlsx'
            except Exception as e:
                print(f"Error loading Excel file: {str(e)}")
                return None, None
        else:
            print(f"Neither {csv_path} nor {xlsx_path} found in {os.getcwd()}")
            return None, None
    
    # Try to load training data (if requested)
    train_df, train_format = None, None
    if train_file_base is not None:
        train_df, train_format = try_load_file(train_file_base)
        if train_df is None and train_file_base != "training_data":  # Only raise error if explicitly requested
            raise FileNotFoundError(f"Could not load training data file with base name '{train_file_base}'")
    
    # Try to load prediction data (optional)
    predict_df, predict_format = None, None
    if predict_file_base is not None:
        predict_df, predict_format = try_load_file(predict_file_base)
    
    return train_df, predict_df, train_format

# -----------------------------------------------------------------------------
# 3) Data Preprocessing
# -----------------------------------------------------------------------------
def preprocess_data(train_df, predict_df=None):
    """
    Prepare the breast cancer dataset(s) for model training and prediction.
    
    For training data:
    1. Drop unnecessary columns (ID, Unnamed: 32)
    2. Encode diagnosis (M/B) to numeric (1/0)
    3. Split features and target
    
    For prediction data (if provided):
    1. Apply same preprocessing steps as training data
    2. Apply same scaling as training data
    
    Arguments:
        train_df: DataFrame with training data (includes diagnosis)
        predict_df: Optional DataFrame with prediction data (without diagnosis)
    
    Returns:
    - X_train: Feature DataFrame for training
    - y_train: Target Series for training
    - X_predict: Feature DataFrame for prediction (or None if not provided)
    - feature_cols: List of feature column names
    - label_encoder: Fitted LabelEncoder for diagnosis
    - scaler: Fitted StandardScaler for features
    """
    # Make copies to avoid modifying the originals
    train_data = train_df.copy()
    
    # Drop non-feature columns in training data
    if 'Unnamed: 32' in train_data.columns:
        train_data = train_data.drop('Unnamed: 32', axis=1)
    
    # Drop ID column if it exists
    if 'id' in train_data.columns:
        train_data = train_data.drop('id', axis=1)
    
    # Encode the diagnosis column
    label_encoder = LabelEncoder()
    train_data['diagnosis_encoded'] = label_encoder.fit_transform(train_data['diagnosis'])
    
    # Print mapping for clarity
    diagnosis_mapping = dict(zip(label_encoder.classes_, label_encoder.transform(label_encoder.classes_)))
    print(f"Diagnosis encoding: {diagnosis_mapping}")
    
    # Get feature columns (all except diagnosis and diagnosis_encoded)
    feature_cols = [col for col in train_data.columns if col not in ['diagnosis', 'diagnosis_encoded']]
    
    # Normalize features (important for this type of medical data)
    scaler = StandardScaler()
    train_data[feature_cols] = scaler.fit_transform(train_data[feature_cols])
    
    # Split into X_train and y_train
    X_train = train_data[feature_cols]
    y_train = train_data['diagnosis_encoded']
    
    # Process prediction data if provided
    X_predict = None
    if predict_df is not None:
        predict_data = predict_df.copy()
        
        # Drop non-feature columns in prediction data
        if 'Unnamed: 32' in predict_data.columns:
            predict_data = predict_data.drop('Unnamed: 32', axis=1)
        
        # Drop ID column if it exists, but keep track of IDs for output
        id_column = None
        if 'id' in predict_data.columns:
            id_column = predict_data['id'].copy()
            predict_data = predict_data.drop('id', axis=1)
        
        # Check for diagnosis column in prediction data (not expected, but handle it)
        if 'diagnosis' in predict_data.columns:
            print("Warning: 'diagnosis' column found in prediction data but will be ignored")
            predict_data = predict_data.drop('diagnosis', axis=1)
        
        # Ensure all feature columns exist in prediction data
        for col in feature_cols:
            if col not in predict_data.columns:
                raise ValueError(f"Feature column '{col}' missing in prediction data")
        
        # Apply the same scaling as training data
        predict_data[feature_cols] = scaler.transform(predict_data[feature_cols])
        
        # Final prediction features
        X_predict = predict_data[feature_cols]
        
        # Reattach ID column if it existed
        if id_column is not None:
            X_predict = X_predict.copy()
            X_predict['id'] = id_column
    
    return X_train, y_train, X_predict, feature_cols, label_encoder, scaler

# -----------------------------------------------------------------------------
# 4) Cross-Validation
# -----------------------------------------------------------------------------
def cross_validate_model(X, y, n_splits=5):
    """
    Performs Stratified K-Fold cross-validation on the dataset using a RandomForestClassifier.
    Returns the array of accuracy scores and prints the mean.
    """
    model = RandomForestClassifier(class_weight='balanced', random_state=42)
    
    skf = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)
    cv_scores = cross_val_score(model, X, y, cv=skf, scoring='accuracy')
    
    print(f"K-Fold CV accuracy scores: {cv_scores}")
    print(f"Mean CV accuracy: {cv_scores.mean():.4f}")
    return cv_scores

# -----------------------------------------------------------------------------
# 5) Hyperparameter Tuning (Grid Search)
# -----------------------------------------------------------------------------
def grid_search_tuning(X, y):
    """
    Performs a Grid Search for hyperparameter tuning on a RandomForestClassifier.
    Returns the best model found.
    """
    model = RandomForestClassifier(random_state=42)
    
    param_grid = {
        'n_estimators': [50, 100, 200],
        'max_depth': [None, 10, 20, 30],
        'min_samples_split': [2, 5, 10],
        'class_weight': ['balanced', None]
    }
    
    grid_search = GridSearchCV(
        estimator=model,
        param_grid=param_grid,
        scoring='f1',  # F1 score is good for binary classification with potential imbalance
        cv=5,
        n_jobs=-1,
        verbose=1
    )
    
    print("Starting Grid Search...")
    grid_search.fit(X, y)
    
    print(f"\nBest Parameters: {grid_search.best_params_}")
    print(f"Best F1 Score: {grid_search.best_score_:.4f}")
    
    best_model = grid_search.best_estimator_
    return best_model

# -----------------------------------------------------------------------------
# 6) Hyperparameter Tuning with Optuna (Bayesian Optimization)
# -----------------------------------------------------------------------------
def tune_rf_with_optuna(X, y, n_trials=30):
    """
    Uses Optuna to perform Bayesian Optimization on a RandomForestClassifier.
    Tries 'n_trials' different hyperparameter combinations, guided by past results.
    Returns the best model found.
    """
    def objective(trial):
        # Suggest hyperparameters
        n_estimators = trial.suggest_int("n_estimators", 50, 300, step=50)
        max_depth = trial.suggest_categorical("max_depth", [None, 10, 20, 30, 40])
        min_samples_split = trial.suggest_int("min_samples_split", 2, 10)
        min_samples_leaf = trial.suggest_int("min_samples_leaf", 1, 5)
        max_features = trial.suggest_categorical("max_features", ['sqrt', 'log2', None])
        class_weight = trial.suggest_categorical("class_weight", ['balanced', None])
        
        # Create the model
        model = RandomForestClassifier(
            n_estimators=n_estimators,
            max_depth=max_depth,
            min_samples_split=min_samples_split,
            min_samples_leaf=min_samples_leaf,
            max_features=max_features,
            class_weight=class_weight,
            random_state=42
        )
        
        # Cross-validation for scoring
        skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
        scores = cross_val_score(model, X, y, cv=skf, scoring='f1', n_jobs=-1)
        return scores.mean()  # maximize F1 score
    
    study = optuna.create_study(direction='maximize')
    study.optimize(objective, n_trials=n_trials)
    
    print(f"\n[Optuna] Number of finished trials: {len(study.trials)}")
    print(f"[Optuna] Best trial parameters: {study.best_trial.params}")
    print(f"[Optuna] Best trial CV F1 score: {study.best_trial.value:.4f}")
    
    # Build final model using the best parameters
    best_params = study.best_params
    best_model = RandomForestClassifier(
        **best_params,
        random_state=42
    )
    best_model.fit(X, y)
    
    return best_model

# -----------------------------------------------------------------------------
# 7) Train a Model
# -----------------------------------------------------------------------------
def train_model(X, y, model=None):
    """
    Trains a Random Forest model or a provided model.
    
    Arguments:
        X: Feature DataFrame
        y: Target Series
        model: Pre-initialized model (optional)
    
    Returns:
        model: Trained model
        X_test: Test features
        y_test: Test target values
    """
    # Split into train and test sets
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.25, random_state=42, stratify=y
    )
    
    # If no model is provided, use a default RandomForest
    if model is None:
        model = RandomForestClassifier(
            n_estimators=100, 
            class_weight='balanced',
            random_state=42
        )
    
    model.fit(X_train, y_train)
    
    train_acc = model.score(X_train, y_train)
    test_acc = model.score(X_test, y_test)
    
    print(f"✅ Model Training Complete!")
    print(f"Training Accuracy: {train_acc:.4f}")
    print(f"Test Accuracy:     {test_acc:.4f}")
    
    return model, X_test, y_test, X_train, y_train

# -----------------------------------------------------------------------------
# 8) Model Evaluation
# -----------------------------------------------------------------------------
def evaluate_model(model, X_test, y_test, label_encoder):
    """
    Evaluates the model with multiple metrics and visualizations.
    
    Arguments:
        model: Trained model
        X_test: Test features
        y_test: Test target values
        label_encoder: LabelEncoder for converting numeric predictions to labels
    
    Returns:
        Dictionary of evaluation metrics
    """
    y_pred = model.predict(X_test)
    y_pred_proba = model.predict_proba(X_test)[:, 1]  # Probability of the positive class
    
    # Convert numeric predictions back to original labels for display
    y_test_labels = label_encoder.inverse_transform(y_test)
    y_pred_labels = label_encoder.inverse_transform(y_pred)
    
    # Print metrics
    print("\nClassification Report:")
    print(classification_report(y_test, y_pred))
    
    accuracy = accuracy_score(y_test, y_pred)
    precision = precision_score(y_test, y_pred)
    recall = recall_score(y_test, y_pred)
    f1 = f1_score(y_test, y_pred)
    
    print(f"Accuracy:  {accuracy:.4f}")
    print(f"Precision: {precision:.4f}")
    print(f"Recall:    {recall:.4f}")
    print(f"F1 Score:  {f1:.4f}")

    # Confusion Matrix
    cm = confusion_matrix(y_test, y_pred)
    plt.figure(figsize=(8, 6))
    disp = ConfusionMatrixDisplay(
        confusion_matrix=cm, 
        display_labels=label_encoder.classes_
    )
    disp.plot(cmap='Blues')
    plt.title("Confusion Matrix")
    plt.tight_layout()
    plt.savefig("confusion_matrix.png")
    print("\nConfusion matrix saved as 'confusion_matrix.png'")
    
    # ROC Curve
    fpr, tpr, _ = roc_curve(y_test, y_pred_proba)
    roc_auc = auc(fpr, tpr)
    
    plt.figure(figsize=(8, 6))
    plt.plot(fpr, tpr, color='darkorange', lw=2, label=f'ROC curve (area = {roc_auc:.2f})')
    plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('Receiver Operating Characteristic (ROC) Curve')
    plt.legend(loc="lower right")
    plt.tight_layout()
    plt.savefig("roc_curve.png")
    print("ROC curve saved as 'roc_curve.png'")
    
    # Feature Importance
    feature_names = X_test.columns
    feature_importance = model.feature_importances_
    
    # Sort features by importance
    sorted_idx = np.argsort(feature_importance)
    
    plt.figure(figsize=(10, 8))
    plt.barh(range(len(sorted_idx)), feature_importance[sorted_idx])
    plt.yticks(range(len(sorted_idx)), [feature_names[i] for i in sorted_idx])
    plt.title('Feature Importance')
    plt.tight_layout()
    plt.savefig("feature_importance.png")
    print("Feature importance plot saved as 'feature_importance.png'")
    
    return {
        "accuracy": accuracy,
        "precision": precision,
        "recall": recall,
        "f1_score": f1,
        "confusion_matrix": cm,
        "roc_auc": roc_auc
    }

# -----------------------------------------------------------------------------
# 9) Feature Analysis
# -----------------------------------------------------------------------------
def analyze_features(X, y, feature_names, label_encoder):
    """
    Analyzes feature distributions and correlations
    
    Arguments:
        X: Feature DataFrame
        y: Target Series
        feature_names: List of feature column names
        label_encoder: LabelEncoder for diagnosis labels
    """
    # Reconstruct dataframe with features and target
    data = X.copy()
    data['diagnosis'] = label_encoder.inverse_transform(y)
    
    # 1. Correlation Heatmap (for mean features only to keep it readable)
    mean_features = [col for col in feature_names if '_mean' in col]
    
    plt.figure(figsize=(12, 10))
    correlation_matrix = data[mean_features].corr()
    sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', fmt='.2f', linewidths=0.5)
    plt.title('Correlation Matrix of Mean Features')
    plt.tight_layout()
    plt.savefig("correlation_matrix.png")
    print("Correlation matrix saved as 'correlation_matrix.png'")
    
    # 2. Feature distributions by diagnosis for the most important features
    # Get the 5 most important mean features
    features_to_plot = [col for col in mean_features if 'mean' in col][:5]
    
    plt.figure(figsize=(15, 10))
    for i, feature in enumerate(features_to_plot):
        plt.subplot(2, 3, i+1)
        sns.boxplot(x='diagnosis', y=feature, data=data)
        plt.title(f'Distribution of {feature}')
        
    plt.tight_layout()
    plt.savefig("feature_distributions.png")
    print("Feature distributions saved as 'feature_distributions.png'")

# -----------------------------------------------------------------------------
# 10) Make Predictions on New Data
# -----------------------------------------------------------------------------
def make_predictions(model, X_predict, label_encoder, output_file='breast_cancer_predictions.xlsx'):
    """
    Makes predictions on new data without diagnosis
    
    Arguments:
        model: Trained model
        X_predict: Feature DataFrame for prediction
        label_encoder: LabelEncoder to convert numeric predictions to labels
        output_file: Path to save predictions Excel file
        
    Returns:
        DataFrame with predictions
    """
    # Check if there's an ID column to preserve
    id_column = None
    if 'id' in X_predict.columns:
        id_column = X_predict['id'].copy()
        X_predict = X_predict.drop('id', axis=1)
    
    # Make predictions (both class and probability)
    y_pred = model.predict(X_predict)
    y_pred_proba = model.predict_proba(X_predict)[:, 1]  # Probability of the positive class
    
    # Convert numeric predictions to labels
    y_pred_labels = label_encoder.inverse_transform(y_pred)
    
    # Create results DataFrame
    results = pd.DataFrame({
        'predicted_diagnosis': y_pred_labels,
        'malignant_probability': y_pred_proba.round(4)
    })
    
    # Add back the ID column if it existed
    if id_column is not None:
        results.insert(0, 'id', id_column)
    
    return results

# -----------------------------------------------------------------------------
# 11) Prediction Pipeline Function with Saved Model
# -----------------------------------------------------------------------------
def predict_with_saved_model(predict_file_base="prediction_data", model_file='breast_cancer_model.pkl'):
    """
    Complete pipeline for making predictions using a saved model:
    1. Load the saved model
    2. Load and preprocess prediction data
    3. Make predictions
    4. Save results to Excel
    
    Arguments:
        predict_file_base: Base name of prediction data file (without extension)
        model_file: Path to the saved model file
        
    Returns:
        success: Boolean indicating if prediction was successful
    """
    print(f"\n=== Making Predictions with Saved Model ===")
    
    # 1. Load the saved model
    model, scaler, label_encoder, feature_names, success = load_saved_model(model_file)
    if not success:
        print(f"Could not load model from '{model_file}'")
        return False
    
    # 2. Load prediction data
    _, predict_df, _ = load_data(None, predict_file_base)
    if predict_df is None:
        print(f"No prediction data found with base name '{predict_file_base}'")
        return False
    
    # 3. Preprocess prediction data
    predict_data = predict_df.copy()
    
    # Handle ID column
    id_column = None
    if 'id' in predict_data.columns:
        id_column = predict_data['id'].copy()
        predict_data = predict_data.drop('id', axis=1)
    
    # Remove unused columns
    if 'Unnamed: 32' in predict_data.columns:
        predict_data = predict_data.drop('Unnamed: 32', axis=1)
    
    if 'diagnosis' in predict_data.columns:
        print("Warning: 'diagnosis' column found in prediction data but will be ignored")
        predict_data = predict_data.drop('diagnosis', axis=1)
    
    # Check features match
    missing_features = [col for col in feature_names if col not in predict_data.columns]
    if missing_features:
        print(f"❌ Error: The following features are missing from prediction data: {missing_features}")
        return False
    
    # Only keep required features in the correct order
    X_predict = predict_data[feature_names].copy()
    
    # Apply scaling
    X_predict_scaled = scaler.transform(X_predict)
    X_predict = pd.DataFrame(X_predict_scaled, columns=feature_names)
    
    # Reattach ID if it existed
    if id_column is not None:
        X_predict['id'] = id_column
    
    # 4. Make predictions
    predictions = make_predictions(model, X_predict, label_encoder)
    
    # 5. Save to Excel
    output_file = 'breast_cancer_predictions.xlsx'
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        predictions.to_excel(writer, sheet_name='Predictions', index=False)
        
        # Format the worksheet
        workbook = writer.book
        worksheet = writer.sheets['Predictions']
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add color coding
        try:
            from openpyxl.styles import PatternFill
            
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            
            pred_col_idx = None
            for idx, col in enumerate(predictions.columns):
                if col == 'predicted_diagnosis':
                    pred_col_idx = idx + 1  # +1 because Excel is 1-indexed
            
            if pred_col_idx:
                for row_idx, value in enumerate(predictions['predicted_diagnosis'], start=2):
                    cell = worksheet.cell(row=row_idx, column=pred_col_idx)
                    if value == 'M':  # Malignant
                        cell.fill = red_fill
                    else:  # Benign
                        cell.fill = green_fill
        except Exception as e:
            print(f"Note: Basic Excel file saved without color formatting: {str(e)}")
    
    print(f"✅ Predictions saved to '{output_file}'")
    print("\nSample predictions:")
    print(predictions.head())
    return True

# -----------------------------------------------------------------------------
# 12) Main Execution - Modified to try saved model first
# -----------------------------------------------------------------------------
if __name__ == "__main__":
    print("=== Breast Cancer Diagnosis Prediction using Random Forest ===")
    
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Breast Cancer Diagnosis Classifier')
    parser.add_argument('--train', help='Base name of training data file (without extension)', default='training_data')
    parser.add_argument('--predict', help='Base name of prediction data file (without extension)', default='prediction_data')
    parser.add_argument('--model', help='Path to saved model file', default='breast_cancer_model.pkl')
    parser.add_argument('--force-train', action='store_true', help='Force training even if saved model exists')
    
    # Check for legacy command line style (for backward compatibility)
    if len(sys.argv) > 1 and not sys.argv[1].startswith('--'):
        # Old style: python script.py train_file predict_file
        train_file_base = sys.argv[1].split('.')[0]
        predict_file_base = sys.argv[2].split('.')[0] if len(sys.argv) > 2 else 'prediction_data'
        args = argparse.Namespace(
            train=train_file_base, 
            predict=predict_file_base, 
            model='breast_cancer_model.pkl', 
            force_train=False
        )
    else:
        # New style with named arguments
        args = parser.parse_args()
    
    print(f"Model file: {args.model}")
    print(f"Prediction data: {args.predict}.csv or {args.predict}.xlsx")
    
    # Try to use saved model first unless force-train is specified
    used_saved_model = False
    if not args.force_train and os.path.exists(args.model):
        print(f"Saved model found at {args.model}, attempting to use it...")
        used_saved_model = predict_with_saved_model(args.predict, args.model)
    
    # If saved model wasn't used (doesn't exist, failed to load, or force-train), train a new one
    if not used_saved_model:
        print("\nProceeding with model training workflow...")
        print(f"Training data: {args.train}.csv or {args.train}.xlsx")
        
        # 1. Load Data
        train_df, predict_df, train_format = load_data(args.train, args.predict)
        
        if train_df is None:
            print(f"❌ Error: Training data not found with base name '{args.train}'")
            sys.exit(1)
            
        # 2. Preprocess Data
        X_train, y_train, X_predict, feature_names, label_encoder, scaler = preprocess_data(train_df, predict_df)
        print(f"Training features shape: {X_train.shape}")
        print(f"Training target shape: {y_train.shape}")
        
        # 3. Feature Analysis
        print("\n=== Feature Analysis ===")
        analyze_features(X_train, y_train, feature_names, label_encoder)
        
        # 4. Baseline Cross-Validation
        print("\n=== Cross-Validation (Baseline RandomForest) ===")
        cross_validate_model(X_train, y_train, n_splits=5)
        
        # 5a. Grid Search
        print("\n=== Hyperparameter Tuning with Grid Search ===")
        best_model_grid = grid_search_tuning(X_train, y_train)
        
        # 5b. Bayesian Optimization (Optuna)
        print("\n=== Hyperparameter Tuning with Optuna (Bayesian Optimization) ===")
        best_model_optuna = tune_rf_with_optuna(X_train, y_train, n_trials=20)
        
        # Compare grid search and optuna models
        print("\n=== Comparing Grid Search and Optuna Models ===")
        skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
        
        grid_scores = cross_val_score(best_model_grid, X_train, y_train, cv=skf, scoring='f1')
        optuna_scores = cross_val_score(best_model_optuna, X_train, y_train, cv=skf, scoring='f1')
        
        print(f"Grid Search model mean F1:   {grid_scores.mean():.4f}")
        print(f"Optuna model mean F1:        {optuna_scores.mean():.4f}")

        # Choose the better model
        if optuna_scores.mean() > grid_scores.mean():
            print("\nOptuna model is better. Using best_model_optuna for final training.")
            chosen_model = best_model_optuna
        else:
            print("\nGrid model is better. Using best_model_grid for final training.")
            chosen_model = best_model_grid
        
        # 6. Train the final model
        print("\n=== Training Final Model with Best Hyperparameters ===")
        final_model, X_test, y_test, X_train_split, y_train_split = train_model(X_train, y_train, model=chosen_model)
        
        # 7. Evaluate the model
        print("\n=== Final Model Evaluation ===")
        eval_results = evaluate_model(final_model, X_test, y_test, label_encoder)
        
        print("\n=== Breast Cancer Diagnosis Model Complete ===")
        print(f"Final model accuracy: {eval_results['accuracy']:.4f}")
        print(f"Final model F1 score: {eval_results['f1_score']:.4f}")
        
        # 8. Save the model and results
        model_file = args.model
        with open(model_file, 'wb') as f:
            pickle.dump({
                'model': final_model,
                'scaler': scaler,
                'label_encoder': label_encoder,
                'feature_names': feature_names
            }, f)
        print(f"\nModel and preprocessing components saved as '{model_file}'")
        
        # Save model evaluation results to Excel
        results_file = 'model_evaluation_results.xlsx'
        
        # Create a DataFrame with evaluation metrics
        metrics_df = pd.DataFrame({
            'Metric': ['Accuracy', 'Precision', 'Recall', 'F1 Score', 'ROC AUC'],
            'Value': [
                eval_results['accuracy'],
                eval_results['precision'],
                eval_results['recall'],
                eval_results['f1_score'],
                eval_results['roc_auc']
            ]
        })
        
        # Save to Excel with formatting
        try:
            with pd.ExcelWriter(results_file, engine='openpyxl') as writer:
                metrics_df.to_excel(writer, sheet_name='Metrics', index=False)
                
                # Format the worksheet
                workbook = writer.book
                worksheet = writer.sheets['Metrics']
                
                # Adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"Model evaluation results saved as '{results_file}'")
        except Exception as e:
            print(f"Warning: Could not save formatted Excel results: {str(e)}")
            # Fallback to CSV
            metrics_df.to_csv('model_evaluation_results.csv', index=False)
            print("Model evaluation results saved as 'model_evaluation_results.csv'")
        
        # 9. Make predictions on new data if provided
        if X_predict is not None:
            print("\n=== Making Predictions on New Data ===")
            predictions = make_predictions(final_model, X_predict, label_encoder)
            
            # Save predictions as Excel file
            output_file = 'breast_cancer_predictions.xlsx'
            
            try:
                # Create Excel writer with formatting
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    predictions.to_excel(writer, sheet_name='Predictions', index=False)
                    
                    # Get the worksheet to apply formatting
                    workbook = writer.book
                    worksheet = writer.sheets['Predictions']
                    
                    # Add some formatting
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Add color coding for malignant/benign predictions if possible
                    try:
                        from openpyxl.styles import PatternFill
                        
                        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                        green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
                        
                        pred_col_idx = None
                        for idx, col in enumerate(predictions.columns):
                            if col == 'predicted_diagnosis':
                                pred_col_idx = idx + 1  # +1 because Excel is 1-indexed
                        
                        if pred_col_idx:
                            for row_idx, value in enumerate(predictions['predicted_diagnosis'], start=2):  # start=2 to skip header
                                cell = worksheet.cell(row=row_idx, column=pred_col_idx)
                                if value == 'M':  # Malignant
                                    cell.fill = red_fill
                                else:  # Benign
                                    cell.fill = green_fill
                    except Exception as e:
                        # If formatting fails, just continue without it
                        print(f"Note: Basic Excel file saved without color formatting: {str(e)}")
                
                print(f"Predictions saved to '{output_file}'")
            except Exception as e:
                print(f"Warning: Could not save formatted Excel predictions: {str(e)}")
                # Fallback to CSV
                predictions.to_csv('breast_cancer_predictions.csv', index=False)
                print("Predictions saved as 'breast_cancer_predictions.csv'")
            
            # Display first few predictions
            print("\nSample predictions:")
            print(predictions.head())